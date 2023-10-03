import csv
import os
import openpyxl

bank_es = 'REGBANESP_CONESTAB_A.xlsx'

if not os.path.isfile(bank_es):
    print('Not found REGBANESP_CONESTAB_A file. Download and convert to XLSx '
        'file (openpyxl support XLSx)')
    print('wget http://www.bde.es/f/webbde/SGE/regis/ficheros/es/REGBANESP_CONESTAB_A.XLS')
    exit()

book = openpyxl.load_workbook(bank_es)
# We filter by 'bancos, cajas de ahorro, cooperativas de credito, entidades
# de credito comunitarias y extracomunitarias'
bank_types = ['BP', 'CA', 'CC', 'CO', 'EFC', 'OR', 'SECC', 'SECE']
bic_codes = {}

with open('bic.csv', 'rt') as bic:
    reader = csv.reader(bic, delimiter=',')
    for code in reader:
        bic_codes[code[0]] = code[2]
# The bank info is in the first sheet
sheet = book.active
headers = [cell.value for cell in sheet[1]]
headers.append('BIC')
with open('bank.csv', 'wt') as f:
    writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore',
            quoting=csv.QUOTE_NONNUMERIC, lineterminator="\n",
            delimiter=',', quotechar='"')
    writer.writeheader()
    for row in sheet.iter_rows():
        # d = {headers[col_index]: sheet.cell(row_index, col_index).value for col_index in range(sheet.ncols)}
        d = {headers[x]: row[x].value for x in range(len(headers)-1)}
        if d['COD_TIPO'] in bank_types and d['FCHBAJA'] in ('', None):
            for value in list(d):
                # Encode all the values to prevent an import crash
                d[value] = d[value] if value is not None else ''
                d['BIC'] = None
                if d['COD_BE'] in bic_codes:
                    d['BIC'] = bic_codes[d['COD_BE']]
            writer.writerow(d)
